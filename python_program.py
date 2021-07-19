"""
This program is checked for pylint
author - Vishwas
date -  15 july 2021
time - 6:30 pm
"""

import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook("Excel.xlsx")
sheets = wb.sheetnames
length = len(sheets)
sheet = wb["SemesterMarks"]
row = sheet.max_row
column = sheet.max_column


def finding(num):
    """
     :param num:
    :return count:
    """
    count = 0
    for j in range(1, row + 1):
        for k in range(1, column + 1):
            case = wb.active.cell(j, k).value
            if num == case:
                count = j
    return count


def checking(word):
    """
    :param word:
    :return value or term:
    """
    term = "sorry"
    for x_num in range(0, length):
        if word == sheets[x_num]:
            return 80
    return term


print("\U0001F642 " + "Welcome " + "\U0001F642")
print()
print("copy your preferred data from one excel file to other excel file")
print()
try:
    print("The ps.numbers in excel are:", end=" ")
    for var in range(2, row + 1):
        print(sheet.cell(var, 1).value, end=',')
    ps_no = int(input("\nChoose one ps.number from them:"))
    print("The sheets in excel are:", end=" ")
    for i in range(0, length):
        print(sheets[i], end=',')
    sh = input("\nChoose one sheet from them:")
    try:
        val = checking(sh)
        check = 80 + val
    except TypeError:
        print("sorry,no sheet with that name")
    mainSheet = wb[sh]
    cellValue = finding(ps_no)
    my_list = []
    wo = Workbook()
    wo['Sheet'].title = "Extracted data"
    sheet_1 = wo.active
    for g in range(1, column + 1):
        variable_1 = mainSheet.cell(1, g).value
        sheet_1.cell(1, g).value = variable_1
    for r in range(1, column + 1):
        variable = mainSheet.cell(cellValue, r).value
        sheet_1.cell(2, r).value = variable
    wo.save("Output_excel.xlsx")
    print("Your preferred data is now available in Output_excel.xlsx file")
except ValueError:
    print("sorry,no such ps number exist")
except KeyError:
    print("wrong data")
